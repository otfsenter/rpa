# coding: utf-8

import openpyxl
from ToolLibrary.base import LibraryComponent, keyword


class FileKeywords(LibraryComponent):

    @keyword
    def read_file(self, path_file):
        content_list = []
        with open(path_file, encoding='utf-8') as f:
            for i in f:
                content_list.append(i.strip())
        return content_list

    @keyword
    def read_excel(self, path_excel, sheet_name):
        # 加载excel文件
        wb = openpyxl.load_workbook(path_excel)
        # 获取当前激活的工作簿
        ws = wb[sheet_name]

        # 获取当前工作簿，已有数据的最大行数、最大列数
        max_row = ws.max_row
        max_col = ws.max_column

        # 读取工作簿里面的内容
        result = []
        for row_index in range(max_row):
            row_index = row_index + 1

            row_list = []
            for col_index in range(max_col):
                col_index = col_index + 1

                cell_value = ws.cell(row=row_index, column=col_index).value
                row_list.append(cell_value)
            result.append(row_list)
        return result
